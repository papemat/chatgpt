import uuid
import csv
import json
import os
from datetime import datetime
from typing import List, Dict
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from zipfile import ZipFile


def export_to_pdf(data: Dict, filepath: str) -> str:
    """
    Esporta un dizionario in PDF con layout chiave: valore su righe separate.
    Usa font Arial o Courier per compatibilità.
    """
    c = canvas.Canvas(filepath, pagesize=A4)
    width, height = A4
    y = height - 40
    c.setFont("Courier", 12)
    for k, v in data.items():
        c.drawString(40, y, f"{k}: {v}")
        y -= 20
        if y < 40:
            c.showPage()
            y = height - 40
            c.setFont("Courier", 12)
    c.save()
    return filepath


def export_to_excel(data: List[Dict], filepath: str) -> str:
    """
    Esporta una lista di dict in Excel con header bold e fill color, colonne auto-adattate.
    """
    wb = Workbook()
    ws = wb.active
    if not data:
        wb.save(filepath)
        return filepath
    headers = list(data[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    for row in data:
        ws.append([row.get(h, "") for h in headers])
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(filepath)
    return filepath


def export_to_csv(data: List[Dict], filepath: str) -> str:
    """
    Esporta una lista di dict in CSV.
    """
    if not data:
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            pass
        return filepath
    headers = list(data[0].keys())
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)
    return filepath


def export_bundle(data: List[Dict], bundle_path: str, readme_text: str = "Export bundle generato da TokIntel API.") -> str:
    """
    Crea un bundle ZIP con export.csv, export.pdf, export.xlsx, README.txt e metadata.json.
    """
    temp_dir = os.path.dirname(bundle_path)
    uid = uuid.uuid4().hex[:8]
    csv_path = os.path.join(temp_dir, f"export_{uid}.csv")
    pdf_path = os.path.join(temp_dir, f"export_{uid}.pdf")
    xlsx_path = os.path.join(temp_dir, f"export_{uid}.xlsx")
    metadata_path = os.path.join(temp_dir, f"metadata_{uid}.json")
    export_to_csv(data, csv_path)
    export_to_excel(data, xlsx_path)
    # Per PDF, prendi solo il primo elemento se lista
    export_to_pdf(data[0] if data else {}, pdf_path)
    with open(metadata_path, 'w', encoding='utf-8') as f:
        json.dump({
            "timestamp": datetime.now().isoformat(),
            "fields": list(data[0].keys()) if data else [],
            "num_records": len(data),
            "files": [os.path.basename(csv_path), os.path.basename(pdf_path), os.path.basename(xlsx_path), "README.txt", os.path.basename(metadata_path)]
        }, f, indent=2)
    with ZipFile(bundle_path, 'w') as zipf:
        zipf.write(csv_path, arcname="export.csv")
        zipf.write(pdf_path, arcname="export.pdf")
        zipf.write(xlsx_path, arcname="export.xlsx")
        zipf.write(metadata_path, arcname="metadata.json")
        readme_path = os.path.join(temp_dir, "README.txt")
        with open(readme_path, 'w', encoding='utf-8') as f:
            f.write(readme_text)
        zipf.write(readme_path, arcname="README.txt")
    # Cleanup temporanei
    for p in [csv_path, pdf_path, xlsx_path, metadata_path, readme_path]:
        try:
            os.remove(p)
        except Exception:
            pass
    return bundle_path 