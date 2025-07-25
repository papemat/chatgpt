"""
Modulo di utility per esportazione avanzata (PDF, CSV, Excel, ZIP) - TokIntel v2 Extensions

Funzioni:
- export_to_pdf(data: dict, filepath: str)
- export_to_csv(data: list, filepath: str)
- export_to_excel(data: list, filepath: str)
- export_to_zip(data: dict, output_dir: str)

Tutte le funzioni sono documentate e testabili in isolamento.
"""

import csv
import os
import zipfile
import json
from typing import Dict, List
from pathlib import Path
from datetime import datetime

# Se vuoi esportare in PDF, assicurati che reportlab sia installato
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Se vuoi esportare in Excel, assicurati che openpyxl sia installato
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def export_to_csv(data: List[dict], filepath: str) -> None:
    """
    Esporta una lista di dizionari in un file CSV.
    :param data: Lista di dizionari (ogni dict è una riga)
    :param filepath: Percorso di destinazione del file CSV
    """
    if not data or data is None:
        raise ValueError("Dati vuoti: impossibile esportare CSV.")
    with open(filepath, mode='w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)


def export_to_pdf(data: Dict, filepath: str) -> None:
    """
    Esporta un dizionario in un file PDF.
    :param data: Dizionario da esportare
    :param filepath: Percorso di destinazione del file PDF
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab non installato. Installa con 'pip install reportlab'.")
    c = canvas.Canvas(filepath, pagesize=letter)
    width, height = letter
    y = height - 40
    c.setFont("Helvetica", 12)
    c.drawString(40, y, "Esportazione dati:")
    y -= 30
    for key, value in data.items():
        c.drawString(40, y, f"{key}: {value}")
        y -= 20
        if y < 40:
            c.showPage()
            y = height - 40
    c.save()


def export_to_excel(data: List[dict], filepath: str) -> None:
    """
    Esporta una lista di dizionari in un file Excel (.xlsx).
    :param data: Lista di dizionari (ogni dict è una riga)
    :param filepath: Percorso di destinazione del file Excel
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl non installato. Installa con 'pip install openpyxl'.")
    
    if not data or data is None:
        raise ValueError("Dati vuoti: impossibile esportare Excel.")
    
    # Crea nuovo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "TokIntel Export"
    
    # Stili per header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Scrivi header
    headers = list(data[0].keys())
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Scrivi dati
    for row, row_data in enumerate(data, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=row_data.get(header, ""))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salva file
    wb.save(filepath)


def export_to_zip(data: Dict, output_dir: str) -> str:
    """
    Esporta tutti i dati in formato ZIP con file multipli.
    :param data: Dizionario contenente tutti i dati da esportare
    :param output_dir: Directory di output per il file ZIP
    :return: Percorso del file ZIP creato
    """
    # Crea directory di output se non esiste
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    # Nome file ZIP con timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_filename = f"tokintel_export_{timestamp}.zip"
    zip_path = os.path.join(output_dir, zip_filename)
    
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # Esporta dati in CSV
        if 'csv_data' in data and data['csv_data']:
            csv_content = []
            for row in data['csv_data']:
                csv_content.append(','.join([f'"{str(v)}"' for v in row.values()]))
            csv_content.insert(0, ','.join([f'"{k}"' for k in data['csv_data'][0].keys()]))
            zipf.writestr('export.csv', '\n'.join(csv_content))
        
        # Esporta dati in JSON
        if 'json_data' in data:
            zipf.writestr('export.json', json.dumps(data['json_data'], indent=2, ensure_ascii=False))
        
        # Esporta metadati
        metadata = {
            'export_date': timestamp,
            'version': '2.1.0',
            'files_included': list(zipf.namelist())
        }
        zipf.writestr('metadata.json', json.dumps(metadata, indent=2))
        
        # Aggiungi README
        readme_content = f"""TokIntel Export - {timestamp}

Questo file ZIP contiene l'esportazione completa dei dati TokIntel.

File inclusi:
- export.csv: Dati in formato CSV
- export.json: Dati in formato JSON
- metadata.json: Metadati dell'esportazione

Generato con TokIntel v2.1.0
"""
        zipf.writestr('README.txt', readme_content)
    
    return zip_path


def export_bundle(data: Dict, output_dir: str = ".") -> str:
    """
    Esporta un bundle completo con tutti i formati disponibili.
    :param data: Dizionario con 'csv_data' e 'json_data'
    :param output_dir: Directory di output
    :return: Percorso del file ZIP bundle
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    bundle_dir = Path(output_dir) / f"tokintel_bundle_{timestamp}"
    bundle_dir.mkdir(parents=True, exist_ok=True)
    
    # Esporta in tutti i formati
    if 'csv_data' in data and data['csv_data']:
        csv_path = bundle_dir / "export.csv"
        export_to_csv(data['csv_data'], str(csv_path))
    
    if 'json_data' in data:
        pdf_path = bundle_dir / "export.pdf"
        export_to_pdf(data['json_data'], str(pdf_path))
        
        if 'csv_data' in data and data['csv_data']:
            excel_path = bundle_dir / "export.xlsx"
            export_to_excel(data['csv_data'], str(excel_path))
    
    # Crea ZIP del bundle
    zip_path = export_to_zip(data, str(bundle_dir.parent))
    
    return zip_path


# TODO: aggiungi test in test_export_tools.py
# TODO: integra in UI con bottoni Streamlit 